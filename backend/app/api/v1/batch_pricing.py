#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Optional, List, Dict
import uuid
import io
import csv
from datetime import datetime
import logging

from app.repository.batch_pricing_repo import BatchPricingRepository
import os

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

router = APIRouter()
logger = logging.getLogger(__name__)


def _normalize_header(header: str) -> str:
    h = (header or '').strip().lower()
    mapping = {
        '物料编码': 'material_code', '编码': 'material_code', '料号': 'material_code', '物料号': 'material_code',
        '物料名称': 'material_name', '名称': 'material_name', '品名': 'material_name',
        '规格': 'specification', '型号': 'specification', '规格型号': 'specification',
        '工艺': 'process_requirements', '要求': 'process_requirements', '工艺要求': 'process_requirements',
        '数量': 'quantity', 'qty': 'quantity', '计量单位': 'uom', '单位': 'uom'
    }
    return mapping.get(h, h)


@router.post("/pricing/batch/upload")
async def upload_batch_pricing(
    file: UploadFile = File(...),
    task_name: Optional[str] = Form(None)
):
    if load_workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请先安装依赖：pip install openpyxl")

    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active

    # 读取表头
    headers = [cell.value if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    normalized_headers = [_normalize_header(str(h)) for h in headers]

    # 读取部分预览行
    preview = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if idx > 20:
            break
        item = {normalized_headers[i]: (row[i] if i < len(row) else None) for i in range(len(normalized_headers))}
        item['row_index'] = idx
        preview.append(item)

    # 估算总行数
    total_rows = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else 0
    trace_id = f"BP-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"

    # 将上传文件保存到本地（与trace关联）
    upload_dir = os.path.join(os.path.dirname(__file__), '../../..', 'uploads')
    upload_dir = os.path.abspath(upload_dir)
    os.makedirs(upload_dir, exist_ok=True)
    saved_path = os.path.join(upload_dir, f"{trace_id}_{file.filename}")
    with open(saved_path, 'wb') as f:
        f.write(content)

    repo = BatchPricingRepository()
    repo.create_task(trace_id, task_name, file.filename, total_rows, normalized_headers, saved_path)

    return {
        "trace_id": trace_id,
        "task_name": task_name,
        "file_name": file.filename,
        "total_rows": total_rows,
        "normalized_columns": normalized_headers,
        "preview_rows": preview
    }


@router.get("/pricing/batch/{trace_id}/preview")
async def get_preview(trace_id: str):
    repo = BatchPricingRepository()
    task = repo.get_task(trace_id)
    if not task:
        raise HTTPException(status_code=404, detail="trace_id 不存在")
    return task


@router.post("/pricing/batch/{trace_id}/run")
async def run_batch_pricing(trace_id: str):
    try:
        # 读取保存的Excel，进行字段映射与简化的价格估算
        saved_path = task.get('source_file_path')
        logger.info(f"[batch-run] trace={trace_id} saved_path={saved_path}")
        if not saved_path or not os.path.exists(saved_path):
            repo.update_task_status(trace_id, 'failed', stats={"error": "源文件缺失"})
            raise HTTPException(status_code=400, detail="源文件缺失，无法执行")

        wb = load_workbook(filename=saved_path, data_only=True)
        ws = wb.active
        headers = [cell.value if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        normalized_headers = [_normalize_header(str(h)) for h in headers]

        def get_val(row, key: str):
            try:
                idx = normalized_headers.index(key)
            except ValueError:
                return None
            return row[idx] if idx < len(row) else None

        results: List[Dict] = []
        success_count = 0
        failed_count = 0
        row_index = 0
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            mc = get_val(row, 'material_code') or get_val(row, '核算物料') or get_val(row, '编码')
            mn = get_val(row, 'material_name') or get_val(row, '核算物料描述') or get_val(row, '名称')
            spec = get_val(row, 'specification') or get_val(row, '规格型号') or get_val(row, '规格')
            proc = get_val(row, 'process_requirements') or get_val(row, '工艺要求') or get_val(row, '工艺')
            qty = get_val(row, 'quantity') or 1
            uom = get_val(row, 'uom') or 'EA'

            if not (mc or mn):
                failed_count += 1
                results.append({
                    'row_index': row_index,
                    'material_code': mc or '',
                    'material_name': mn or '',
                    'specification': spec or '',
                    'process_requirements': proc or '',
                    'quantity': qty or 1,
                    'uom': uom,
                    'estimated_price': None,
                    'currency': 'CNY',
                    'status': 'failed',
                    'reason_or_notes': '缺少物料编码或名称',
                    'rule_version': 'v1.0'
                })
                continue

            # 简化估价：基于字段长度/数量的启发式（占位，可替换为真实核价逻辑）
            base = 100.0
            name_factor = len(str(mn or mc)) * 0.5
            spec_factor = len(str(spec or '')) * 0.2
            proc_factor = len(str(proc or '')) * 0.1
            price = round((base + name_factor + spec_factor + proc_factor) * float(qty or 1), 2)

            success_count += 1
            results.append({
                'row_index': row_index,
                'material_code': mc or '',
                'material_name': mn or '',
                'specification': spec or '',
                'process_requirements': proc or '',
                'quantity': qty or 1,
                'uom': uom,
                'estimated_price': price,
                'currency': 'CNY',
                'status': 'success',
                'reason_or_notes': '',
                'rule_version': 'v1.0'
            })

            if len(results) % 500 == 0:
                repo.insert_results(trace_id, results)
                results = []

        if results:
            repo.insert_results(trace_id, results)

        repo.update_task_status(trace_id, 'completed', stats={
            'success_count': success_count,
            'failed_count': failed_count,
            'total_processed': row_index
        })

        return {"success": True, "success_count": success_count, "failed_count": failed_count}
    except HTTPException as he:
        logger.exception(f"[batch-run] HTTP错误: {he.detail}")
        repo.update_task_status(trace_id, 'failed', stats={"error": str(he.detail)})
        return {"success": False, "error": str(he.detail)}
    except Exception as e:
        # 避免纯文本500，统一返回JSON，便于前端展示
        try:
            logger.exception(f"[batch-run] 执行失败: {e}")
        except Exception:
            print('[batch-run] 执行失败:', e)
        repo.update_task_status(trace_id, 'failed', stats={"error": str(e)})
        return {"success": False, "error": str(e)}


@router.get("/pricing/batch/{trace_id}/results")
async def list_results(trace_id: str, status: Optional[str] = 'all', pn: int = 1, ps: int = 50):
    repo = BatchPricingRepository()
    return repo.list_results(trace_id, status, pn, ps)


@router.get("/pricing/batch/{trace_id}/export")
async def export_results(trace_id: str, format: str = 'csv'):
    repo = BatchPricingRepository()
    data = repo.list_results(trace_id, 'all', 1, 100000)

    if format == 'csv':
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(data['rows'][0].keys()) if data['rows'] else [])
        if data['rows']:
            writer.writeheader()
            for r in data['rows']:
                writer.writerow(r)
        buf.seek(0)
        return StreamingResponse(iter([buf.getvalue()]), media_type='text/csv', headers={
            'Content-Disposition': f'attachment; filename="{trace_id}.csv"'
        })
    else:
        return JSONResponse(content=data)


@router.post("/pricing/batch/{trace_id}/approve")
async def approve_results(trace_id: str, approve: bool = True, approver: Optional[str] = None):
    repo = BatchPricingRepository()
    task = repo.get_task(trace_id)
    if not task:
        raise HTTPException(status_code=404, detail="trace_id 不存在")
    repo.approve_task(trace_id, approver or 'leader', approve)
    return {"success": True, "status": 'approved' if approve else 'rejected'}


