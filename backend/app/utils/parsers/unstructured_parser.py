from typing import Tuple
import io
import pdfplumber
from openpyxl import load_workbook

# 说明：PDF/Excel解析为纯文本的最小实现


def parse_pdf_to_text(content: bytes) -> str:
    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
    return "\n".join(text_parts).strip()


def parse_excel_to_text(content: bytes) -> str:
    buf = io.BytesIO(content)
    wb = load_workbook(buf, read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            line = ",".join([str(c) if c is not None else "" for c in row])
            parts.append(line)
    return "\n".join(parts).strip()


def detect_and_parse(filename: str, content: bytes) -> Tuple[str, str]:
    name = filename.lower()
    if name.endswith('.pdf'):
        return ('pdf', parse_pdf_to_text(content))
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        return ('excel', parse_excel_to_text(content))
    # 其他类型默认按文本
    try:
        return ('text', content.decode('utf-8'))
    except Exception:
        return ('binary', '')


